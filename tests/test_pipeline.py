from pathlib import Path

from openpyxl import load_workbook

from app.config import Settings
from app.jobs import create_analysis_job, list_analysis_jobs, run_analysis_job
from app.pipeline import ingest_knowledge, process_tariffs
from app.reporting import apply_approved_updates, load_audit_log
from app.review_store import save_review_decision
from app.schemas import ReviewDecision


def make_settings(tmp_path: Path) -> Settings:
    return Settings(
        INPUT_EXCEL_PATH=Path("data/input/tariff_packs.xlsx"),
        OUTPUT_DIR=tmp_path / "output",
        CHROMA_DIR=tmp_path / "chroma",
        GEMINI_API_KEY="",
    )


def test_pipeline_generates_preview_proposals_without_llm(tmp_path: Path, monkeypatch) -> None:
    settings = make_settings(tmp_path)

    def fail_if_called(*args, **kwargs):
        raise AssertionError("Preview mode should not call Gemini")

    monkeypatch.setattr("app.agent.GeminiProvider.propose_update", fail_if_called)
    ingest_result = ingest_knowledge(settings)
    result = process_tariffs(settings)

    assert ingest_result["source_counts"] == {"confluence": 10, "wiki": 7, "email": 8}
    assert result.mode == "preview"
    assert len(result.records) == 28
    assert any(
        proposal.pack_id == "PK001"
        and proposal.field_name == "price_azn"
        and proposal.proposed_value == 12.9
        and proposal.source_conflict_detected
        for proposal in result.proposals
    )


def test_gemini_mode_falls_back_without_api_key(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    result = process_tariffs(settings, generation_mode="gemini")

    assert result.mode == "fallback"
    assert result.proposals


def test_analysis_job_tracks_completed_run(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    job = create_analysis_job("preview", settings)

    completed = run_analysis_job(job.job_id, settings)
    runs = list_analysis_jobs(settings)

    assert completed.status == "completed"
    assert completed.progress == 100
    assert completed.stage == "Completed"
    assert completed.actual_mode == "preview"
    assert completed.summary is not None
    assert completed.summary.records == 28
    assert runs[0].job_id == job.job_id


def test_apply_approved_updates_only_mutates_approved_fields(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    process_tariffs(settings)
    save_review_decision(
        ReviewDecision(
            pack_id="PK001",
            field_name="price_azn",
            decision="approved",
            reviewer="pytest",
            reasoning="Approved sample price correction.",
        ),
        settings,
    )
    save_review_decision(
        ReviewDecision(
            pack_id="PK001",
            field_name="activation_code",
            decision="rejected",
            reviewer="pytest",
            reasoning="Rejected sample activation code correction.",
        ),
        settings,
    )

    result = apply_approved_updates(settings)
    workbook = load_workbook(result["output_excel"], read_only=True, data_only=True)
    sheet = workbook["tariff_packs"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    price_idx = headers.index("price_azn")
    activation_idx = headers.index("activation_code")
    row = next(row for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] == "PK001")

    assert row[price_idx] == 12.9
    assert row[activation_idx] == "*123*09#"
    assert Path(result["audit_log"]).exists()
    assert Path(result["report"]).exists()
    assert {entry["decision"] for entry in load_audit_log(settings)} == {"approved", "rejected"}
